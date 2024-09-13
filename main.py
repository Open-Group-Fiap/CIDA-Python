from typing import Dict, List
from azure.ai.inference.models import SystemMessage, UserMessage
from azure.core.credentials import AzureKeyCredential
from fastapi import FastAPI, Response, status
from azure.storage.blob import BlobClient, BlobServiceClient
from pydantic import BaseModel
from io import BytesIO
import google.generativeai as genai
import os
import dotenv
import requests
from azure.ai.inference import ChatCompletionsClient

dotenv.load_dotenv()
client = inference_client = ChatCompletionsClient(
    endpoint=os.getenv("AZURE_AI_ENDPOINT") or "",
    credential=AzureKeyCredential(os.getenv("AZURE_AI_API_KEY") or ""),
)
app = FastAPI()

account_url = "https://cidastore.blob.core.windows.net"

url_API = "https://api.cidainstitute.com"


class AnalyzeRequest(BaseModel):
    container: str
    file_names: List[str]
    id_usuario: int
    ids_arquivos: List[int]


class AnalyzeResponse(BaseModel):
    pass


def process_file(file: bytes, file_name: str):
    file_data = BytesIO(file)
    text = ""
    if file_name.endswith(".pdf"):
        text = extract_text_from_pdf(file_data)
    elif file_name.endswith(".docx") or file_name.endswith(".doc"):
        text = extract_text_from_docx(file_data)
    elif file_name.endswith(".xlsx"):
        text = extract_text_from_xlsx(file_data)
    elif file_name.endswith(".csv"):
        text = extract_text_from_csv(file_data)
    elif file_name.endswith(".txt"):
        text = file_data.read().decode("utf-8")
    else:
        raise ValueError("Unsupported file type")
    return text


def extract_text_from_pdf(file_data: BytesIO):
    from PyPDF2 import PdfReader

    pdfReader = PdfReader(file_data)
    text = ""
    for page in pdfReader.pages:
        text += page.extract_text()
    return text


def extract_text_from_docx(file_data: BytesIO):
    from docx2txt import process

    document = process(file_data)
    return document


def extract_text_from_xlsx(file_data: BytesIO):
    from openpyxl import load_workbook

    loader = load_workbook(file_data)
    text = ""
    for sheet in loader.worksheets:
        for row in sheet.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text


def extract_text_from_csv(file_data: BytesIO):
    from csv import reader

    text = ""
    for row in reader(file_data.read().decode("utf-8").splitlines()):
        text += " ".join(row) + "\n"
    return text


def get_blob(container_name, blob_name):
    blob_service_client = BlobServiceClient(account_url=account_url)
    container_client = blob_service_client.get_container_client(container_name)
    blob_client = container_client.get_blob_client(blob_name)
    return blob_client


async def post_resumo(descricao_resumo: str, id_usuario: int):
    response = requests.post(
        f"{url_API}/resumo",
        json={"descricao": descricao_resumo, "idUsuario": id_usuario},
    )

    resumo = response.json()

    return resumo["IdResumo"]


async def put_arquivo(id_resumo: int, ids_arquivos: List[int]):
    for i in ids_arquivos:
        requests.put(
            f"{url_API}/arquivo/{i}",
            json={"idResumo": id_resumo},
        )


async def post_insight(descricao_insight: str, id_usuario: int, id_resumo: int):
    response = requests.post(
        f"{url_API}/insight",
        json={
            "descricao": descricao_insight,
            "idUsuario": id_usuario,
            "idResumo": id_resumo,
        },
    )

    return response.status_code


@app.get("/")
def read_root():
    return {"Hello": "World"}


@app.post("/analyze")
async def analyze(data: AnalyzeRequest, responseStatus: Response) -> AnalyzeResponse:
    blobs: List[BlobClient] = []
    text = ""
    for file_name in data.file_names:
        blob = get_blob(data.container, file_name)
        blobs.append(blob)
    for blob in blobs:
        try:
            converted = blob.download_blob().readall()
            text += process_file(converted, blob.blob_name) + "\n"
        except Exception as e:
            print(blob.url)
            print(e)
    start_phase = "Resuma o seguinte documento e diminua seu tamanho total, mantenha a coesão, os dados e as estatisticas: "
    response = client.complete(
        messages=[SystemMessage(content=start_phase), UserMessage(content=text)]
    )
    text = response.choices[0].message.content

    # Post do resumo
    id_resumo = await post_resumo(
        text, idUsuario=data.id_usuario, ids_arquivos=data.ids_arquivos
    )
    # put do id_resumo nos arquivos
    await put_arquivo(id_resumo, data.ids_arquivos)

    gemini_prompt = "Você é a CIDA, Consulting Insights With Deep Analysis, você tem como função analisar relatorios empresariais e ajudar as empresas gerando insights, analises e recomendações com base nos dados apresentados. Com base nesse relatorio empresarial, analise os dados e a situação da empresa e gere insights destacando os dados mais relevantes, pontos fortes e pontos mais fracos:"
    gemini_prompt += "\n\n"
    gemini_prompt += text
    genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
    model = genai.GenerativeModel("gemini-1.5-flash-latest")
    response = model.generate_content(gemini_prompt)

    # Post do insight
    status_code = await post_insight(
        response.text, id_usuario=data.id_usuario, id_resumo=id
    )

    if status_code == 201:
        responseStatus.status_code = status.HTTP_200_OK
    else:
        responseStatus.status_code = status.HTTP_500_INTERNAL_SERVER_ERROR

    return AnalyzeResponse()


@app.get("/get-all-blobs/{container}")
async def get_all_blobs(container: str) -> Dict[str, str]:
    blob_service_client = BlobServiceClient(account_url=account_url)
    container_client = blob_service_client.get_container_client(container)
    blobs = container_client.list_blobs()
    blobs_dict = {}
    for blob in blobs:
        blobs_dict[blob.name] = get_blob(container, blob.name).url
    return blobs_dict


if __name__ == "__main__":
    print("De preferência, use o comando abaixo para iniciar o servidor:")
    print("python -m uvicorn main:app --reload")
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
